#!/usr/bin/env python3
"""
ЭКСПОРТ ВСЕХ ТОВАРОВ В EXCEL
Выгружает ВСЕ 4052 товара из БД для ручной обработки
"""

import os

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

print("🔍 Загружаю ВСЕ товары из БД...")

# Получаем все категории
categories = supabase.table("categories").select("id, name, slug").execute()
cat_dict = {cat["id"]: cat for cat in categories.data}

print(f"📦 Категорий: {len(categories.data)}")

# Получаем ВСЕ товары (с пагинацией)
all_products = []
offset = 0
batch_size = 1000

while True:
    products = (
        supabase.table("products")
        .select("*")
        .range(offset, offset + batch_size - 1)
        .execute()
    )

    if not products.data:
        break

    all_products.extend(products.data)
    offset += batch_size
    print(f"  Загружено: {len(all_products)} товаров...")

    if len(products.data) < batch_size:
        break

print(f"📊 Всего загружено: {len(all_products)} товаров")

# Создаём Excel файл
wb = Workbook()
ws = wb.active
ws.title = "All Products"

# Заголовки
headers = [
    "ID",
    "Название товара",
    "Цена",
    "Старая цена",
    "В наличии",
    "Категория",
    "Slug категории",
    "Артикул",
    "Бренд (категория)",
    "Тип запчасти",
    "Описание",
]

# Форматируем заголовки
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Добавляем данные
universal_count = 0
brand_count = 0

for row_num, product in enumerate(all_products, 2):
    category = cat_dict.get(product["category_id"], {})
    cat_slug = category.get("slug", "")
    cat_name = category.get("name", "")

    # Определяем бренд и тип из slug категории
    # Формат: brand-type или universal-type
    brand = ""
    part_type = ""

    if cat_slug:
        parts = cat_slug.split("-", 1)
        if len(parts) >= 1:
            brand = parts[0].upper()
        if len(parts) == 2:
            part_type = parts[1].replace("-", " ").title()

    # Подсветка
    is_universal = "universal" in cat_slug
    if is_universal:
        universal_count += 1
    else:
        brand_count += 1

    # Записываем данные
    ws.cell(row=row_num, column=1).value = product["id"]
    ws.cell(row=row_num, column=2).value = product["name"]
    ws.cell(row=row_num, column=3).value = product.get("price")
    ws.cell(row=row_num, column=4).value = product.get("old_price")
    ws.cell(row=row_num, column=5).value = (
        "ДА" if product.get("in_stock", False) else "НЕТ"
    )
    ws.cell(row=row_num, column=6).value = cat_name
    ws.cell(row=row_num, column=7).value = cat_slug
    ws.cell(row=row_num, column=8).value = product.get("sku", "")
    ws.cell(row=row_num, column=9).value = brand
    ws.cell(row=row_num, column=10).value = part_type
    ws.cell(row=row_num, column=11).value = product.get("description", "")

    # Подсвечиваем Universal зелёным, бренды - голубым
    if is_universal:
        fill_color = "E2EFDA"  # Зелёный
    else:
        fill_color = "DEEBF7"  # Голубой

    for col in range(1, 12):
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )

# Настраиваем ширину колонок
column_widths = {
    1: 8,  # ID
    2: 60,  # Название
    3: 10,  # Цена
    4: 10,  # Старая цена
    5: 12,  # В наличии
    6: 35,  # Категория
    7: 30,  # Slug
    8: 15,  # Артикул
    9: 20,  # Бренд
    10: 25,  # Тип
    11: 40,  # Описание
    11: 40,  # Описание
}

for col, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Включаем автофильтр
ws.auto_filter.ref = f"A1:K{len(all_products) + 1}"

# Замораживаем первую строку
ws.freeze_panes = "A2"

# Сохраняем файл
excel_file = "/home/ibm/dongfeng-minitraktor/ALL_PRODUCTS.xlsx"
wb.save(excel_file)

print(f"\n✅ Excel файл создан: {excel_file}")
print(f"📊 Экспортировано товаров: {len(all_products)}")
print()
print(f"🟢 Universal товаров:    {universal_count} (подсвечены зелёным)")
print(f"🔵 Брендовых товаров:    {brand_count} (подсвечены голубым)")
print()
print("📝 Колонки:")
print("  1. ID")
print("  2. Название товара")
print("  3. Цена")
print("  4. Старая цена")
print("  5. В наличии (ДА/НЕТ)")
print("  6. Категория")
print("  7. Slug категории")
print("  8. Артикул")
print("  9. Бренд (из категории)")
print("  10. Тип запчасти")
print("  11. Описание")
print()
print("✨ Особенности:")
print("  ✓ Автофильтры на всех колонках")
print("  ✓ Universal товары - ЗЕЛЁНАЯ подсветка")
print("  ✓ Брендовые товары - ГОЛУБАЯ подсветка")
print("  ✓ Первая строка закреплена")
print()
print("🎯 Готово к фильтрации и сортировке!")
