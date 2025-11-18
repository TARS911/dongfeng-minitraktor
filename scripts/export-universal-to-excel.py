#!/usr/bin/env python3
"""
ЭКСПОРТ UNIVERSAL ТОВАРОВ В EXCEL
Выгружает все товары из Universal категорий в XLSX с фильтрами и форматированием
"""

import os

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

print("🔍 Загружаю товары из Universal категорий...")

# Получаем все Universal категории
categories = (
    supabase.table("categories")
    .select("id, name, slug")
    .like("slug", "universal-%")
    .execute()
)
cat_dict = {cat["id"]: cat for cat in categories.data}
universal_cat_ids = [cat["id"] for cat in categories.data]

print(f"📦 Найдено {len(universal_cat_ids)} Universal категорий")

# Получаем все товары из Universal (без лимита)
all_products = []
offset = 0
batch_size = 1000

while True:
    products = (
        supabase.table("products")
        .select("*")
        .in_("category_id", universal_cat_ids)
        .range(offset, offset + batch_size - 1)
        .execute()
    )

    if not products.data:
        break

    all_products.extend(products.data)
    offset += batch_size
    print(f"  Загружено: {len(all_products)} товаров...")

    if len(products.data) < batch_size:
        break

print(f"📊 Всего загружено: {len(all_products)} товаров")

# Создаём Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Universal Products"

# Заголовки
headers = [
    "ID",
    "Название товара",
    "Цена",
    "Старая цена",
    "В наличии",
    "Категория",
    "Slug категории",
    "Артикул",
    "Бренд (определён)",
    "Описание",
]

# Форматируем заголовки
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Добавляем данные
for row_num, product in enumerate(all_products, 2):
    category = cat_dict.get(product["category_id"], {})

    # Определяем бренд из названия
    name_lower = product["name"].lower()
    detected_brand = ""

    brands_to_check = {
        "S1100": ["s1100", "с1100", "s-1100"],
        "S195": ["s195", "с195", "s-195"],
        "ZS": ["zs1100", "zs1105", "zs1110", "zs1115", "zs1125", "zs195"],
        "R175": ["r175", "р175", "r-175"],
        "R180": ["r180", "р180", "r-180"],
        "DONGFENG": ["dongfeng", "донгфенг", "дунгфенг"],
        "URALETS": ["уралец"],
        "KM": ["км-385", "км385", "км-"],
        "JINMA": ["джинма", "jinma"],
        "FOTON": ["фотон", "foton", "lovol"],
        "XINGTAI": ["синтай", "xingtai"],
        "SHIFENG": ["шифенг", "shifeng"],
        "YTO": ["yto", "юто"],
        "MTZ": ["мтз", "беларус"],
    }

    for brand, patterns in brands_to_check.items():
        for pattern in patterns:
            if pattern in name_lower:
                detected_brand = brand
                break
        if detected_brand:
            break

    # Записываем данные
    ws.cell(row=row_num, column=1).value = product["id"]
    ws.cell(row=row_num, column=2).value = product["name"]
    ws.cell(row=row_num, column=3).value = product.get("price")
    ws.cell(row=row_num, column=4).value = product.get("old_price")
    ws.cell(row=row_num, column=5).value = (
        "ДА" if product.get("in_stock", False) else "НЕТ"
    )
    ws.cell(row=row_num, column=6).value = category.get("name", "")
    ws.cell(row=row_num, column=7).value = category.get("slug", "")
    ws.cell(row=row_num, column=8).value = product.get("sku", "")
    ws.cell(row=row_num, column=9).value = detected_brand
    ws.cell(row=row_num, column=10).value = product.get("description", "")

    # Подсвечиваем товары с определённым брендом
    if detected_brand:
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )

# Настраиваем ширину колонок
column_widths = {
    1: 8,  # ID
    2: 60,  # Название
    3: 10,  # Цена
    4: 10,  # Старая цена
    5: 12,  # В наличии
    6: 35,  # Категория
    7: 30,  # Slug
    8: 15,  # Артикул
    9: 15,  # Бренд
    10: 40,  # Описание
}

for col, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Включаем автофильтр
ws.auto_filter.ref = f"A1:J{len(all_products) + 1}"

# Замораживаем первую строку
ws.freeze_panes = "A2"

# Сохраняем файл
excel_file = "/home/ibm/dongfeng-minitraktor/UNIVERSAL_PRODUCTS.xlsx"
wb.save(excel_file)

print(f"\n✅ Excel файл создан: {excel_file}")
print(f"📊 Экспортировано товаров: {len(all_products)}")
print("\n📝 Колонки в файле:")
print("  1. ID - ID товара в базе")
print("  2. Название товара")
print("  3. Цена")
print("  4. Старая цена")
print("  5. В наличии (ДА/НЕТ)")
print("  6. Категория")
print("  7. Slug категории")
print("  8. Артикул")
print("  9. Бренд (определён) - автоматически найденный бренд")
print("  10. Описание")
print("\n✨ Особенности:")
print("  ✓ Заголовки с форматированием (синий фон)")
print("  ✓ Автофильтры на всех колонках")
print("  ✓ Товары с определённым брендом подсвечены жёлтым")
print("  ✓ Первая строка закреплена при прокрутке")
print("  ✓ Ширина колонок настроена автоматически")
print("\n🎯 Готово к фильтрации и сортировке в Excel!")
